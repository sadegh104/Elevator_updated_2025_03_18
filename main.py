import math
import streamlit as st
from openpyxl import load_workbook
from bs4 import BeautifulSoup
st.set_page_config("Elevator Calculator", layout="wide")
wb = load_workbook('Data-Ranjbar.xlsx')
with st.sidebar:
    st.title("تحلیل محاسبات آسانسور")
cols = st.columns(3)
with cols[0]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb18 = ws['B18']
            cellc18 = ws['C18']
            celld18 = ws['D18']
            new_valueb18 = st.selectbox('نوع موتورخانه', options=[cellc18.value, celld18.value],
                                       key=f'{cellb18.coordinate}_{sheet_name}')
            if new_valueb18 == cellc18.value:
                cellb18.value = cellc18.value
            elif new_valueb18 == celld18.value:
                cellb18.value = celld18.value
with cols[1]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb18 = ws['B18']
            cellc18 = ws['C18']
            celld18 = ws['D18']
            cellb19 = ws['B19']
            cellc19 = ws['C19']
            celld19 = ws['D19']
            if cellb18.value == cellc18.value:
                new_valueb19 = st.selectbox('نوع موتور', options=[cellc19.value, celld19.value],
                                            key=f'{cellb19.coordinate}_{sheet_name}')
                if new_valueb19 == cellc19.value:
                    cellb19.value = cellc19.value
                elif new_valueb19 == celld19.value:
                    cellb19.value = celld19.value
            elif cellb18.value == celld18.value:
                new_valueb19 = st.selectbox('نوع موتور', options=[celld19.value],
                                            key=f'{cellb19.coordinate}_{sheet_name}')
                if new_valueb19 == celld19.value:
                    cellb19.value = celld19.value
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb19 = ws['B19']
            cellc19 = ws['C19']
            celld19 = ws['D19']
            cellb20 = ws['B20']
            cellc20 = ws['C20']
            celld20 = ws['D20']
            celle20 = ws['E20']
            if cellb19.value == cellc19.value:
                new_valueb20 = st.selectbox('سیستم تعلیق', options=[cellc20.value],
                                            key=f'{cellb20.coordinate}_{sheet_name}')
                if new_valueb20 == cellc20.value:
                    cellb20.value = cellc20.value
            elif cellb19.value == celld19.value:
                new_valueb20 = st.selectbox('سیستم تعلیق', options=[celld20.value],
                                            key=f'{cellb20.coordinate}_{sheet_name}')
                if new_valueb20 == cellc20.value:
                    cellb20.value = cellc20.value
                elif new_valueb20 == celld20.value:
                    cellb20.value = celld20.value
                elif new_valueb20 == celle20.value:
                    cellb20.value = celle20.value
cols = st.columns(3)
with cols[0]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb2 = ws['B2']
            new_valueb2 = st.number_input('تعداد آسانسور',
                                          key=f'{cellb2.coordinate}_{sheet_name}', value=1)
with cols[1]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb3 = ws['B3']
            cellc3 = ws['C3']
            celld3 = ws['D3']
            celle3 = ws['E3']
            cellf3 = ws['F3']
            cellg3 = ws['G3']
            cellh3 = ws['H3']
            celli3 = ws['I3']
            cellj3 = ws['J3']
            cellk3 = ws['K3']
            celll3 = ws['L3']
            new_valueb3 = st.selectbox('ظرفیت (نفر)', options=[cellc3.value, celld3.value, celle3.value, cellf3.value, cellg3.value, cellh3.value, celli3.value, cellj3.value, cellk3.value, celll3.value],
                                       key=f'{cellb3.coordinate}_{sheet_name}')
            if new_valueb3 == cellc3.value:
                cellb3.value = cellc3.value
            elif new_valueb3 == celld3.value:
                cellb3.value = celld3.value
            elif new_valueb3 == celle3.value:
                cellb3.value = celle3.value
            elif new_valueb3 == cellf3.value:
                cellb3.value = cellf3.value
            elif new_valueb3 == cellg3.value:
                cellb3.value = cellg3.value
            elif new_valueb3 == cellh3.value:
                cellb3.value = cellh3.value
            elif new_valueb3 == celli3.value:
                cellb3.value = celli3.value
            elif new_valueb3 == cellj3.value:
                cellb3.value = cellj3.value
            elif new_valueb3 == cellk3.value:
                cellb3.value = cellk3.value
            elif new_valueb3 == celll3.value:
                cellb3.value = celll3.value
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            # cellb3 = ws['B3']
            # cellc3 = ws['C3']
            # celld3 = ws['D3']
            # celle3 = ws['E3']
            # cellf3 = ws['F3']
            # cellg3 = ws['G3']
            # cellh3 = ws['H3']
            # celli3 = ws['I3']
            # cellj3 = ws['J3']
            # cellk3 = ws['K3']
            # celll3 = ws['L3']
            cellb4 = ws['B4']
            cellc4 = ws['C4']
            celld4 = ws['D4']
            cellb18 = ws['B18']
            cellc18 = ws['C18']
            celld18 = ws['D18']
            if cellb18.value == cellc18.value:
                new_valueb4 = st.selectbox('نوع آسانسور', options=[celld4.value],
                                           key=f'{cellb4.coordinate}_{sheet_name}')
                if new_valueb4 == celld4.value:
                    cellb4.value = celld4.value
            elif cellb18.value == celld18.value:
                new_valueb4 = st.selectbox('نوع آسانسور', options=[cellc4.value],
                                           key=f'{cellb4.coordinate}_{sheet_name}')
                if new_valueb4 == cellc4.value:
                    cellb4.value = cellc4.value
            # if cellb3.value == cellc3.value:
            #     new_valueb4 = st.selectbox('نوع آسانسور', options=[cellc4.value],
            #                                key=f'{cellb4.coordinate}_{sheet_name}')
            #     if new_valueb4 == cellc4.value:
            #         cellb4.value = cellc4.value
            # elif cellb3.value == celld3.value:
            #     new_valueb4 = st.selectbox('نوع آسانسور', options=[cellc4.value, celld4.value],
            #                                key=f'{cellb4.coordinate}_{sheet_name}')
            #     if new_valueb4 == cellc4.value:
            #         cellb4.value = cellc4.value
            #     elif new_valueb4 == celld4.value:
            #         cellb4.value = celld4.value
            # elif cellb3.value == celle3.value or cellb3.value == cellf3.value:
            #     new_valueb4 = st.selectbox('نوع آسانسور', options=[celld4.value],
            #                                key=f'{cellb4.coordinate}_{sheet_name}')
            #     if new_valueb4 == celld4.value:
            #         cellb4.value = celld4.value
            # new_valueb4 = st.selectbox('نوع آسانسور', options=[celld4.value],
            #                             key=f'{cellb4.coordinate}_{sheet_name}')
            # if new_valueb4 == celld4.value:
            #     cellb4.value = celld4.value
cols = st.columns(3)
with cols[0]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb10 = ws['B10']
            new_valueb10 = st.number_input('(متر) چاهک',
                                           key=f'{cellb10.coordinate}_{sheet_name}', value=1.5)
with cols[1]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb6 = ws['B6']
            cellc6 = ws['C6']
            celld6 = ws['D6']
            celle6 = ws['E6']
            cellf6 = ws['F6']
            cellg6 = ws['G6']
            cellh6 = ws['H6']
            celli6 = ws['I6']
            new_valueb6 = st.selectbox('تعداد توقف',
                                       options=[cellc6.value, celld6.value, celle6.value, cellf6.value, cellg6.value,
                                                cellh6.value, celli6.value],
                                       key=f'{cellb6.coordinate}_{sheet_name}')
            if new_valueb6 == cellc6.value:
                cellb6.value = cellc6.value
            elif new_valueb6 == celld6.value:
                cellb6.value = celld6.value
            elif new_valueb6 == celle6.value:
                cellb6.value = celle6.value
            elif new_valueb6 == cellf6.value:
                cellb6.value = cellf6.value
            elif new_valueb6 == cellg6.value:
                cellb6.value = cellg6.value
            elif new_valueb6 == cellh6.value:
                cellb6.value = cellh6.value
            elif new_valueb6 == celli6.value:
                cellb6.value = celli6.value
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb14 = ws['B14']
            cellc14 = ws['C14']
            celld14 = ws['D14']
            celle14 = ws['E14']
            new_valueb14 = st.selectbox('عرض وزنه تعادل', options=[cellc14.value, celld14.value, celle14.value],
                                        key=f'{cellb14.coordinate}_{sheet_name}')
            if new_valueb14 == cellc14.value:
                cellb14.value = cellc14.value
            elif new_valueb14 == celld14.value:
                cellb14.value = celld14.value
            elif new_valueb14 == celle14.value:
                cellb14.value = celle14.value
cols = st.columns(3)
with cols[0]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb9 = ws['B9']
            new_valueb9 = st.number_input('(متر) اورهد',
                                          key=f'{cellb9.coordinate}_{sheet_name}', value=3.9)
with cols[1]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb11 = ws['B11']
            cellc11 = ws['C11']
            celld11 = ws['D11']
            new_valueb11 = st.selectbox('نوع درب', options=[cellc11.value, celld11.value],
                                        key=f'{cellb11.coordinate}_{sheet_name}')
            if new_valueb11 == cellc11.value:
                cellb11.value = cellc11.value
            elif new_valueb11 == celld11.value:
                cellb11.value = celld11.value
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb11 = ws['B11']
            cellc11 = ws['C11']
            celld11 = ws['D11']
            cellb13 = ws['B13']
            cellc13 = ws['C13']
            celld13 = ws['D13']
            celle13 = ws['E13']
            if cellb11.value == cellc11.value:
                new_valueb13 = st.selectbox('جهت درب', options=[cellc13.value, celld13.value, celle13.value],
                                            key=f'{cellb13.coordinate}_{sheet_name}')
                if new_valueb13 == cellc13.value:
                    cellb13.value = cellc13.value
                elif new_valueb13 == celld13.value:
                    cellb13.value = celld13.value
                elif new_valueb13 == celle13.value:
                    cellb13.value = celle13.value
            elif cellb11.value == celld11.value:
                new_valueb13 = st.selectbox('جهت درب', options=[cellc13.value, celld13.value],
                                            key=f'{cellb13.coordinate}_{sheet_name}')
                if new_valueb13 == cellc13.value:
                    cellb13.value = cellc13.value
                elif new_valueb13 == celld13.value:
                    cellb13.value = celld13.value
cols = st.columns(3)
with cols[0]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb3 = ws['B3']
            cellc3 = ws['C3']
            celld3 = ws['D3']
            celle3 = ws['E3']
            cellf3 = ws['F3']
            cellg3 = ws['G3']
            cellh3 = ws['H3']
            celli3 = ws['I3']
            cellj3 = ws['J3']
            cellk3 = ws['K3']
            celll3 = ws['L3']
            cellb16 = ws['B16']
            if cellb3.value == cellc3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=400)
            elif cellb3.value == celld3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=450)
            elif cellb3.value == celle3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=500)
            elif cellb3.value == cellf3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=550)
            elif cellb3.value == cellg3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=600)
            elif cellb3.value == cellh3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=675)
            elif cellb3.value == celli3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=750)
            elif cellb3.value == cellj3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=800)
            elif cellb3.value == cellk3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=850)
            elif cellb3.value == celll3.value: 
                new_valueb16 = st.number_input('وزن کابین (کیلوگرم)',
                                          key=f'{cellb16.coordinate}_{sheet_name}', value=900)
with cols[1]:
    cellb11 = ws['B11']
    cellc11 = ws['C11']
    celld11 = ws['D11']
    cellb12 = ws['B12']
    cellc12 = ws['C12']
    celld12 = ws['D12']
    celle12 = ws['E12']
    cellf12 = ws['F12']
    if cellb11.value == cellc11.value:
        new_valueb12 = st.selectbox('عرض درب',
                                    options=[cellc12.value, celld12.value, celle12.value, cellf12.value],
                                    key=f'{cellb12.coordinate}_{sheet_name}')
        if new_valueb12 == cellc12.value:
            cellb12.value = cellc12.value
        elif new_valueb12 == celld12.value:
            cellb12.value = celld12.value
        elif new_valueb12 == celle12.value:
            cellb12.value = celle12.value
        elif new_valueb12 == cellf12.value:
            cellb12.value = cellf12.value
    elif cellb11.value == celld11.value:
        new_valueb12 = st.selectbox('عرض درب', options=[cellc12.value, celld12.value, celle12.value],
                                    key=f'{cellb12.coordinate}_{sheet_name}')
        if new_valueb12 == cellc12.value:
            cellb12.value = cellc12.value
        elif new_valueb12 == celld12.value:
            cellb12.value = celld12.value
        elif new_valueb12 == celle12.value:
            cellb12.value = celle12.value
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb5 = ws['B5']
            cellc5 = ws['C5']
            celld5 = ws['D5']
            new_valueb5 = st.selectbox('سرعت آسانسور',
                                       options=[cellc5.value, celld5.value],
                                       key=f'{cellb5.coordinate}_{sheet_name}')
            if new_valueb5 == cellc5.value:
                cellb5.value = cellc5.value
            elif new_valueb5 == celld5.value:
                cellb5.value = celld5.value
cols = st.columns(3)
with cols[2]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb20 = ws['B20']
            cellc20 = ws['C20']
            celld20 = ws['D20']
            celle20 = ws['E20']
            cellb21 = ws['B21']
            if cellb19.value == celld19.value:
                new_valueb21 = st.number_input('فاصله فلکه های کابین (متر)',
                                           key=f'{cellb21.coordinate}_{sheet_name}', value=1.50, min_value=0.00, max_value=2.00)
            else:
                pass
with cols[1]:
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == 'Data (1)':
            ws = wb[sheet_name]
            cellb3 = ws['B3']
            cellc3 = ws['C3']
            celld3 = ws['D3']
            celle3 = ws['E3']
            cellf3 = ws['F3']
            cellg3 = ws['G3']
            cellh3 = ws['H3']
            celli3 = ws['I3']
            cellj3 = ws['J3']
            cellk3 = ws['K3']
            celll3 = ws['L3']
            cellb5 = ws['B5']
            cellc5 = ws['C5']
            celld5 = ws['D5']
            cellb17 = ws['B17']
            cellc17 = ws['C17']
            celld17 = ws['D17']
            celle17 = ws['E17']
            cellf17 = ws['F17']
            cellb19 = ws['B19']
            cellc19 = ws['C19']
            celld19 = ws['D19']
            if cellb5.value == cellc5.value and cellb19.value == cellc19.value and (cellb3.value == cellc3.value or cellb3.value == celld3.value or cellb3.value == celle3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[cellc17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == cellc17.value:
                    cellb17.value = cellc17.value
            elif cellb5.value == cellc5.value and cellb19.value == cellc19.value and (cellb3.value == cellf3.value or cellb3.value == cellg3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celld17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celld17.value:
                    cellb17.value = celld17.value
            elif cellb5.value == cellc5.value and cellb19.value == cellc19.value and (cellb3.value == cellh3.value or cellb3.value == celli3.value or cellb3.value == cellj3.value or cellb3.value == cellk3.value or cellb3.value == celll3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celle17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celle17.value:
                    cellb17.value = celle17.value
            elif cellb5.value == celld5.value and cellb19.value == cellc19.value and (cellb3.value == cellc3.value or cellb3.value == celld3.value or cellb3.value == celle3.value or cellb3.value == cellf3.value or cellb3.value == cellg3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celld17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celld17.value:
                    cellb17.value = celld17.value
            elif cellb5.value == celld5.value and cellb19.value == cellc19.value and (cellb3.value == cellh3.value or cellb3.value == celli3.value or cellb3.value == cellj3.value or cellb3.value == cellk3.value or cellb3.value == celll3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celle17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celle17.value:
                    cellb17.value = celle17.value
            elif cellb5.value == cellc5.value and cellb19.value == celld19.value and (cellb3.value == cellc3.value or cellb3.value == celld3.value or cellb3.value == celle3.value or cellb3.value == cellf3.value or cellb3.value == cellg3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celld17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celld17.value:
                    cellb17.value = celld17.value
            elif cellb5.value == cellc5.value and cellb19.value == celld19.value and (cellb3.value == cellh3.value or cellb3.value == celli3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celle17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celle17.value:
                    cellb17.value = celle17.value
            elif cellb5.value == cellc5.value and cellb19.value == celld19.value and (cellb3.value == cellj3.value or cellb3.value == cellk3.value or cellb3.value == celll3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[cellf17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == cellf17.value:
                    cellb17.value = cellf17.value
            elif cellb5.value == celld5.value and cellb19.value == celld19.value and (cellb3.value == cellc3.value or cellb3.value == celld3.value or cellb3.value == celle3.value or cellb3.value == cellf3.value or cellb3.value == cellg3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celld17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celld17.value:
                    cellb17.value = celld17.value
            elif cellb19.value == celld19.value and (cellb3.value == cellh3.value or cellb3.value == celli3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[celle17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == celle17.value:
                    cellb17.value = celle17.value
            elif cellb19.value == celld19.value and (cellb3.value == cellj3.value or cellb3.value == cellk3.value or cellb3.value == celll3.value):
                new_valueb17 = st.selectbox('تعداد شیار فلکه موتور',
                                    options=[cellf17.value],
                                    key=f'{cellb17.coordinate}_{sheet_name}')
                if new_valueb17 == cellf17.value:
                    cellb17.value = cellf17.value    
                # if cellb3.value == cellc3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=5)
                # elif cellb3.value == celld3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=5)
                # elif cellb3.value == celle3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=5)
                # elif cellb3.value == cellf3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=5)
                # elif cellb3.value == cellg3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=5)
                # elif cellb3.value == cellh3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=6)
                # elif cellb3.value == celli3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=6)
                # elif cellb3.value == cellj3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=6)
                # elif cellb3.value == cellk3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=6)
                # elif cellb3.value == celll3.value: 
                #     new_valueb17 = st.number_input('تعداد شیار فلکه موتور',
                #                               key=f'{cellb17.coordinate}_{sheet_name}', value=6)
if st.button('ثبت'):
    ws1 = wb['Data (1)']
    ws2 = wb['Data (2)']
    ws3 = wb['Data (3)']
    new_valueb3 = ws1['B3'].value
    new_valuec3 = ws1['C3'].value
    new_valued3 = ws1['D3'].value
    new_valuee3 = ws1['E3'].value
    new_valuef3 = ws1['F3'].value
    new_valueg3 = ws1['G3'].value
    new_valueh3 = ws1['H3'].value
    new_valuei3 = ws1['I3'].value
    new_valuej3 = ws1['J3'].value
    new_valuek3 = ws1['K3'].value
    new_valuel3 = ws1['L3'].value
    new_valueb5 = ws1['B5'].value
    new_valuec5 = ws1['C5'].value
    new_valued5 = ws1['D5'].value
    new_valueb8 = (new_valueb6 - 1) * 3.4
    new_valueb11 = ws1['B11'].value
    new_valuec11 = ws1['C11'].value
    new_valued11 = ws1['D11'].value
    new_valueb12 = ws1['B12'].value
    new_valuec12 = ws1['C12'].value
    new_valued12 = ws1['D12'].value
    new_valuee12 = ws1['E12'].value
    new_valuef12 = ws1['F12'].value
    new_valueb13 = ws1['B13'].value
    new_valuec13 = ws1['C13'].value
    new_valued13 = ws1['D13'].value
    new_valuee13 = ws1['E13'].value
    new_valueb6 = ws1['B6'].value
    new_valuec6 = ws1['C6'].value
    new_valued6 = ws1['D6'].value
    new_valuee6 = ws1['E6'].value
    new_valuef6 = ws1['F6'].value
    new_valueg6 = ws1['G6'].value
    new_valueh6 = ws1['H6'].value
    new_valuei6 = ws1['I6'].value
    new_valueb14 = ws1['B14'].value
    new_valuec14 = ws1['C14'].value
    new_valued14 = ws1['D14'].value
    new_valuee14 = ws1['E14'].value
    new_valueb18 = ws1['B18'].value
    new_valuec18 = ws1['C18'].value
    new_valued18 = ws1['D18'].value
    new_valueb19 = ws1['B19'].value
    new_valuec19 = ws1['C19'].value
    new_valued19 = ws1['D19'].value
    new_valueb20 = ws1['B20'].value
    new_valuec20 = ws1['C20'].value
    new_valued20 = ws1['D20'].value
    new_valuee20 = ws1['E20'].value
    EquipmentDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;font-size:1.8rem;"><b>شرح متریال</b></td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;font-size:0.8rem;"><b>تعداد</b></td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;font-size:0.8rem;"><b>قیمت واحد به ریال</b></td></tr></table>
                    """
    st.markdown(EquipmentDescription, unsafe_allow_html=True)
    if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A2'].value}<br>{ws2['A19'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A2'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A3'].value}<br>{ws2['A20'].value}<</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A3'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A4'].value}<br>{ws2['A21'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A4'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A5'].value}<br>{ws2['A22'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A5'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A6'].value}<br>{ws2['A23'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A6'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A7'].value}<br>{ws2['A24'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A7'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A8'].value}<br>{ws2['A25'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A8'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A9'].value}<br>{ws2['A26'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A9'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A10'].value}<br>{ws2['A27'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A10'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A11'].value}<br>{ws2['A28'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A11'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A12'].value}<br>{ws2['A29'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A12'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A13'].value}<br>{ws2['A30'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A13'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A14'].value}<br>{ws2['A31'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A14'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A15'].value}<br>{ws2['A32'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A15'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A16'].value}<br>{ws2['A33'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A16'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            MotorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['A17'].value}<br>{ws2['A34'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['A17'].value:,.0f}</td></tr></table>
                    """
            st.markdown(MotorDescription, unsafe_allow_html=True)
    if ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
        if (((new_valueb8 + new_valueb9 + new_valueb10)*0.4) - int((new_valueb8 + new_valueb9 + new_valueb10)*0.4) > 0.1):    
            RailT90Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['B2'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*(int((new_valueb8 + new_valueb9 + new_valueb10)*0.4)+1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['B2'].value:,.0f}</td></tr></table>
                    """
            st.markdown(RailT90Description, unsafe_allow_html=True)
        else:
            RailT90Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['B2'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*(int((new_valueb8 + new_valueb9 + new_valueb10)*0.4))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['B2'].value:,.0f}</td></tr></table>
                    """
            st.markdown(RailT90Description, unsafe_allow_html=True)
    if ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
        if new_valueb3 == new_valuej3 or new_valueb3 == new_valuek3 or new_valueb3 == new_valuel3:
            if (((new_valueb8 + new_valueb9 + new_valueb10)*0.4) - int((new_valueb8 + new_valueb9 + new_valueb10)*0.4) > 0.1):
                RailT90Quant = (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)
                if (RailT90Quant / new_valueb2) % 2 == 0:
                    PoshtBandRailT90Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT90Quant / new_valueb2) - 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C2'].value:,.0f}</td></tr></table>
                            """
                    st.markdown(PoshtBandRailT90Description, unsafe_allow_html=True)
                else:
                    PoshtBandRailT90Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT90Quant / new_valueb2) - 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C2'].value:,.0f}</td></tr></table>
                            """
                    st.markdown(PoshtBandRailT90Description, unsafe_allow_html=True)
                # PoshtBandRailT90Quantity = f"""
                #                                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)}</div>
                #                                     """
                # st.markdown(PoshtBandRailT90Quantity, unsafe_allow_html=True)
            else:
                RailT90Quant = (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4))
                if (RailT90Quant / new_valueb2) % 2 == 0:
                    PoshtBandRailT90Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT90Quant / new_valueb2) - 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C2'].value:,.0f}</td></tr></table>
                            """
                    st.markdown(PoshtBandRailT90Description, unsafe_allow_html=True)
                else:
                    PoshtBandRailT90Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*((RailT90Quant / new_valueb2) - 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C2'].value:,.0f}</td></tr></table>
                            """
                    st.markdown(PoshtBandRailT90Description, unsafe_allow_html=True)
                # PoshtBandRailT90Quantity = f"""
                #                                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4))}</div>
                #                                     """
                # st.markdown(PoshtBandRailT90Quantity, unsafe_allow_html=True)
    if ws1['B3'].value != ws1['J3'].value and ws1['B3'].value != ws1['K3'].value and ws1['B3'].value != ws1['L3'].value:
        if (((new_valueb8 + new_valueb9 + new_valueb10)*0.4) - int((new_valueb8 + new_valueb9 + new_valueb10)*0.4) > 0.1):
            RailT70Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['B3'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*(int((new_valueb8 + new_valueb9 + new_valueb10)*0.4)+1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['B3'].value:,.0f}</td></tr></table>
                    """
            st.markdown(RailT70Description, unsafe_allow_html=True)
        else:
            RailT70Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['B3'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*(int((new_valueb8 + new_valueb9 + new_valueb10)*0.4))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['B3'].value:,.0f}</td></tr></table>
                    """
            st.markdown(RailT70Description, unsafe_allow_html=True)
    if ws1['B3'].value != ws1['J3'].value and ws1['B3'].value != ws1['K3'].value and ws1['B3'].value != ws1['L3'].value:
        if (((new_valueb8 + new_valueb9 + new_valueb10)*0.4) - int((new_valueb8 + new_valueb9 + new_valueb10)*0.4) > 0.1):
            RailT70Quant = (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)
            if (RailT70Quant / new_valueb2) % 2 == 0:
                PoshtBandRailT70Description = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT70Quant / new_valueb2) - 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C3'].value:,.0f}</td></tr></table>
                        """
                st.markdown(PoshtBandRailT70Description, unsafe_allow_html=True)
            else:
                PoshtBandRailT70Description = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT70Quant / new_valueb2) - 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C3'].value:,.0f}</td></tr></table>
                        """
                st.markdown(PoshtBandRailT70Description, unsafe_allow_html=True)
                # PoshtBandRailT70Quantity = f"""
                #                                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)}</div>
                #                                     """
                # st.markdown(PoshtBandRailT70Quantity, unsafe_allow_html=True)
        else:
            RailT70Quant = (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4))
            if (RailT70Quant / new_valueb2) % 2 == 0:
                PoshtBandRailT70Description = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT70Quant / new_valueb2) - 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C3'].value:,.0f}</td></tr></table>
                        """
                st.markdown(PoshtBandRailT70Description, unsafe_allow_html=True)
            else:
                PoshtBandRailT70Description = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT70Quant / new_valueb2) - 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C3'].value:,.0f}</td></tr></table>
                        """
                st.markdown(PoshtBandRailT70Description, unsafe_allow_html=True)
                # PoshtBandRailT70Quantity = f"""
                #                                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4))}</div>
                #                                     """
                # st.markdown(PoshtBandRailT70Quantity, unsafe_allow_html=True)
    RailT50Description = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['B4'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['B4'].value:,.0f}</td></tr></table>
                                """
    st.markdown(RailT50Description, unsafe_allow_html=True)
    RailT50Quant = (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)
    if (RailT50Quant / new_valueb2) % 2 == 0:
        PoshtBandRailT50Description = f"""
                                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C4'].value}</td>
                                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT50Quant / new_valueb2) - 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C4'].value:,.0f}</td></tr></table>
                                            """
        st.markdown(PoshtBandRailT50Description, unsafe_allow_html=True)
    else:
        PoshtBandRailT50Description = f"""
                                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['C4'].value}</td>
                                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*int((RailT50Quant / new_valueb2) - 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['C4'].value:,.0f}</td></tr></table>
                                            """
        st.markdown(PoshtBandRailT50Description, unsafe_allow_html=True)
        # PoshtBandRailT50Quantity = f"""
        #                                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * (int((new_valueb8 + new_valueb9 + new_valueb10) * 0.4) + 1)}</div>
        #                                     """
        # st.markdown(PoshtBandRailT50Quantity, unsafe_allow_html=True)
    if ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
        LoghmeRailT90Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['D2'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * ((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4)))*2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['D2'].value:,.0f}</td></tr></table>
                """
        st.markdown(LoghmeRailT90Description, unsafe_allow_html=True)
    if ws1['B3'].value != ws1['J3'].value and ws1['B3'].value != ws1['K3'].value and ws1['B3'].value != ws1['L3'].value:
        LoghmeRailT70Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['D3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * ((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4)))*2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['D3'].value:,.0f}</td></tr></table>
                """
        st.markdown(LoghmeRailT70Description, unsafe_allow_html=True)
    LoghmeRailT50Description = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['D4'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * ((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['D4'].value:,.0f}</td></tr></table>
                                """
    st.markdown(LoghmeRailT50Description, unsafe_allow_html=True)
    BracketCabinDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['E2'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * ((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4)))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['E2'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(BracketCabinDescription, unsafe_allow_html=True)
    LeftBracketWazneDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['F2'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * (((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) + 2)))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['F2'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(LeftBracketWazneDescription, unsafe_allow_html=True)
    RightBracketWazneDescription = f"""
                                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['F3'].value}</td>
                                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(math.ceil((new_valueb2) * (((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) + 2)))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['F3'].value:,.0f}</td></tr></table>
                                            """
    st.markdown(RightBracketWazneDescription, unsafe_allow_html=True)
    LeftBracketWazneQuant = math.ceil((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) + 2))
    # LoghmeRailT50Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
    # Pitch10Quantity = f"""
    #                 <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * int(2 * int(LeftBracketWazneQuant) + LoghmeRailT50Quant)}</div>
    #                 """
    # st.markdown(Pitch10Quantity, unsafe_allow_html=True)
    # LoghmeRailT50Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
    Pitch10Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['G2'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(2 * int(LeftBracketWazneQuant))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['G2'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(Pitch10Description, unsafe_allow_html=True)
    if ws1['B3'].value != ws1['J3'].value and ws1['B3'].value != ws1['K3'].value and ws1['B3'].value != ws1['L3'].value:
        BracketCabinQuant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4)))
        # LoghmeRailT70Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
        # Pitch12Quantity = f"""
        #                         <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * int(2 * int(BracketCabinQuant) + LoghmeRailT70Quant)}</div>
        #                         """
        # st.markdown(Pitch12Quantity, unsafe_allow_html=True)
        # LoghmeRailT70Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
        Pitch12Description = f"""
                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['G3'].value}</td>
                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(2 * int(BracketCabinQuant))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['G3'].value:,.0f}</td></tr></table>
                                                """
        st.markdown(Pitch12Description, unsafe_allow_html=True)
    LeftBracketWazneQuant = math.ceil((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) + 2))
    # LoghmeRailT50Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
    # Pitch8Quantity = f"""
    #                         <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2) * int(2 * int(LeftBracketWazneQuant) + LoghmeRailT50Quant)}</div>
    #                         """
    # st.markdown(Pitch8Quantity, unsafe_allow_html=True)
    # LoghmeRailT50Quant = int(math.ceil(((((new_valueb8 + new_valueb9 + new_valueb10) / 1.7) * 2) + 4))) * 2
    Pitch8Description = f"""
                                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['G4'].value}</td>
                                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(2 * int(LeftBracketWazneQuant))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['G4'].value:,.0f}</td></tr></table>
                                            """
    st.markdown(Pitch8Description, unsafe_allow_html=True)
    LasticDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['H2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 4}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['H2'].value:,.0f}</td></tr></table>
                                """
    st.markdown(LasticDescription, unsafe_allow_html=True)
    if new_valueb19 == new_valuec19:    
        if new_valueb20 == new_valuec20:
            if new_valueb3 == new_valuec3 or new_valueb3 == new_valued3 or new_valueb3 == new_valuee3:
                # SimBoxelScore10Quantity = f"""
                # <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(int(new_valueb2))*int(round((new_valueb8 + new_valueb9 + 5) * 4))}</div>
                # """
                # st.markdown(SimBoxelScore10Quantity, unsafe_allow_html=True)
                SimBoxelScore10Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I2'].value:,.0f}</td></tr></table>
                            """
                st.markdown(SimBoxelScore10Description, unsafe_allow_html=True)
            elif new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
                SimBoxelScore10Description = f"""
                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I2'].value}</td>
                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I2'].value:,.0f}</td></tr></table>
                                        """
                st.markdown(SimBoxelScore10Description, unsafe_allow_html=True)
        elif new_valueb20 == new_valued20 or new_valueb20 == new_valuee20:
            if new_valueb3 == new_valuec3 or new_valueb3 == new_valued3 or new_valueb3 == new_valuee3:
                # SimBoxelScore10Quantity = f"""
                # <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(int(new_valueb2))*int(round((new_valueb8 + new_valueb9 + 5) * 4))}</div>
                # """
                # st.markdown(SimBoxelScore10Quantity, unsafe_allow_html=True)
                SimBoxelScore10Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I2'].value:,.0f}</td></tr></table>
                            """
                st.markdown(SimBoxelScore10Description, unsafe_allow_html=True)
            elif new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
                SimBoxelScore10Description = f"""
                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I2'].value}</td>
                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I2'].value:,.0f}</td></tr></table>
                                        """
                st.markdown(SimBoxelScore10Description, unsafe_allow_html=True)
    elif new_valueb19 == new_valued19:
        if new_valueb20 == new_valuec20:
            if new_valueb3 == new_valuec3 or new_valueb3 == new_valued3 or new_valueb3 == new_valuee3:
                # SimBoxelScore10Quantity = f"""
                # <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(int(new_valueb2))*int(round((new_valueb8 + new_valueb9 + 5) * 4))}</div>
                # """
                # st.markdown(SimBoxelScore10Quantity, unsafe_allow_html=True)
                SimBoxelScore8Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                            """
                st.markdown(SimBoxelScore8Description, unsafe_allow_html=True)
            elif new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
                SimBoxelScore8Description = f"""
                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                                        """
                st.markdown(SimBoxelScore8Description, unsafe_allow_html=True)
        elif new_valueb20 == new_valued20 or new_valueb20 == new_valuee20:
            if new_valueb3 == new_valuec3 or new_valueb3 == new_valued3 or new_valueb3 == new_valuee3:
                # SimBoxelScore10Quantity = f"""
                # <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(int(new_valueb2))*int(round((new_valueb8 + new_valueb9 + 5) * 4))}</div>
                # """
                # st.markdown(SimBoxelScore10Quantity, unsafe_allow_html=True)
                SimBoxelScore8Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                            """
                st.markdown(SimBoxelScore8Description, unsafe_allow_html=True)
            elif new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
                SimBoxelScore8Description = f"""
                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                                        """
                st.markdown(SimBoxelScore8Description, unsafe_allow_html=True)
    # cols = st.columns(2)
    # with cols[0]:
    #     if ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         SimBoxelScore11Description = f"""
    #                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;direction:rtl;'>{ws2['I5'].value}</div>
    #                     """
    #         st.markdown(SimBoxelScore11Description, unsafe_allow_html=True)
    # with cols[1]:
    #     if new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
    #         SimBoxelScore11Quantity = f"""
    #             <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * 6))}</div>
    #             """
    #         st.markdown(SimBoxelScore11Quantity, unsafe_allow_html=True)
    if new_valueb19 == new_valuec19:
        if new_valueb20 == new_valuec20:    
            if new_valueb3 == new_valueh3 or new_valueb3 == new_valuei3 or new_valueb3 == new_valuej3 or new_valueb3 == new_valuek3 or new_valueb3 == new_valuel3:
                SimBoxelScore12Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I3'].value:,.0f}</td></tr></table>
                """
                st.markdown(SimBoxelScore12Description, unsafe_allow_html=True)
        elif new_valueb20 == new_valued20 or new_valueb20 == new_valuee20:
            if new_valueb3 == new_valueh3 or new_valueb3 == new_valuei3 or new_valueb3 == new_valuej3 or new_valueb3 == new_valuek3 or new_valueb3 == new_valuel3:
                SimBoxelScore12Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I3'].value:,.0f}</td></tr></table>
                """
                st.markdown(SimBoxelScore12Description, unsafe_allow_html=True)
    elif new_valueb19 == new_valued19:
        if new_valueb20 == new_valuec20:    
            if new_valueb3 == new_valueh3 or new_valueb3 == new_valuei3 or new_valueb3 == new_valuej3 or new_valueb3 == new_valuek3 or new_valueb3 == new_valuel3:
                SimBoxelScore12Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                """
                st.markdown(SimBoxelScore12Description, unsafe_allow_html=True)
        elif new_valueb20 == new_valued20 or new_valueb20 == new_valuee20:
            if new_valueb3 == new_valueh3 or new_valueb3 == new_valuei3 or new_valueb3 == new_valuej3 or new_valueb3 == new_valuek3 or new_valueb3 == new_valuel3:
                SimBoxelScore12Description = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I6'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(round((((new_valueb8 + new_valueb9 + new_valueb10) * 2) + 8 + new_valueb21) * new_valueb17))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I6'].value:,.0f}</td></tr></table>
                """
                st.markdown(SimBoxelScore12Description, unsafe_allow_html=True)
    SimBoxelScore6Description = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['I4'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(math.ceil((new_valueb10 + new_valueb9 + new_valueb8 + 2))) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['I4'].value:,.0f}</td></tr></table>
                                """
    st.markdown(SimBoxelScore6Description, unsafe_allow_html=True)
    if ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value or ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
            GholabBoxelScore10Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['J2'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int((new_valueb17) * 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['J2'].value:,.0f}</td></tr></table>
                    """
            st.markdown(GholabBoxelScore10Description, unsafe_allow_html=True)
    elif ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value or ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
            GholabBoxelScore8Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['J4'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int((new_valueb17) * 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['J4'].value:,.0f}</td></tr></table>
                    """
            st.markdown(GholabBoxelScore8Description, unsafe_allow_html=True)
    if ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            GholabBoxelScore13Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['J3'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int((new_valueb17) * 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['J3'].value:,.0f}</td></tr></table>
                    """
            st.markdown(GholabBoxelScore13Description, unsafe_allow_html=True)
    elif ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            GholabBoxelScore8Description = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['J4'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int((new_valueb17) * 2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['J4'].value:,.0f}</td></tr></table>
                    """
            st.markdown(GholabBoxelScore8Description, unsafe_allow_html=True)
    # if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K2'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K2'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K3'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K3'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K4'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K4'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K5'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K5'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K6'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K6'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K7'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K7'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K8'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K8'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K9'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K9'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K10'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K10'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K11'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K11'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K12'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K12'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K13'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K13'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K14'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K14'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K15'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K15'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K16'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K16'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         FalakeHarzgardDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['K17'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['K17'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(FalakeHarzgardDescription, unsafe_allow_html=True)
    # if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ2'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ2'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ3'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ3'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ4'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ4'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ5'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ5'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ6'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ6'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ7'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ7'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ8'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ8'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ9'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ9'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ10'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ10'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ11'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ11'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ12'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ12'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ13'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ13'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ14'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ14'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ15'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ15'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ16'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ16'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         BoxelDiameterDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AZ1'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{ws2['AZ17'].value}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AZ17'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(BoxelDiameterDescription, unsafe_allow_html=True)
    if new_valueb5 == new_valuec5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valuec19: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L2'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L2'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valued5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valuec19: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L3'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valuec5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valued19 and new_valueb20 == new_valuec20: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L4'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L4'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valuec5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valued19 and new_valueb20 != new_valuec20: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L6'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L6'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valued5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valued19 and new_valueb20 == new_valuec20: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L5'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L5'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valued5 and new_valueb18 == new_valuec18 and new_valueb19 == new_valued19 and new_valueb20 != new_valuec20: 
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L7'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L7'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valuec5 and new_valueb18 == new_valued18:
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L8'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L8'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    elif new_valueb5 == new_valued5 and new_valueb18 == new_valued18:
        GovernorUpDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['L9'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['L9'].value:,.0f}</td></tr></table>
                        """
        st.markdown(GovernorUpDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B13'].value == ws1['C13'].value:
        LeftDoorGhoflDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['M2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['M2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(LeftDoorGhoflDescription, unsafe_allow_html=True)
    elif ws1['B11'].value == ws1['D11'].value and ws1['B13'].value == ws1['D13'].value:
        RightDoorGhoflDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['M3'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['M3'].value:,.0f}</td></tr></table>
                                """
        st.markdown(RightDoorGhoflDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B13'].value == ws1['C13'].value:
        DictatorDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['N2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['N2'].value:,.0f}</td></tr></table>
                            """
        st.markdown(DictatorDescription, unsafe_allow_html=True)
    elif ws1['B11'].value == ws1['D11'].value and ws1['B13'].value == ws1['D13'].value:
        DictatorDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['N2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['N2'].value:,.0f}</td></tr></table>
                            """
        st.markdown(DictatorDescription, unsafe_allow_html=True)
    DoorKeyDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['O2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['O2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(DoorKeyDescription, unsafe_allow_html=True)
    if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
        PolyortanDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['P2'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['P2'].value:,.0f}</td></tr></table>
                """
        st.markdown(PolyortanDescription, unsafe_allow_html=True)
    elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
        PolyortanDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['P3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['P3'].value:,.0f}</td></tr></table>
                """
        st.markdown(PolyortanDescription, unsafe_allow_html=True)
    elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
        BufferDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Q2'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Q2'].value:,.0f}</td></tr></table>
                        """
        st.markdown(BufferDescription, unsafe_allow_html=True)
    RoghandanDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['R2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 4}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['R2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(RoghandanDescription, unsafe_allow_html=True)
    AshkiDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['S2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['S2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(AshkiDescription, unsafe_allow_html=True)
    CorpiScore6Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 6}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(CorpiScore6Description, unsafe_allow_html=True)
    if ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            # SimBoxelScore10Quant = (int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * 4))
            # SimBoxelScore10Quant = (int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * 5))
            # SimBoxelScore10Quant = (int(new_valueb2)) * int(round(((new_valueb8 + 9) - 2) * 6))
            # CorpiScore10Quantity = f"""
            #                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{(SimBoxelScore10Quant) * 4}</div>
            #                     """
            # st.markdown(CorpiScore10Quantity, unsafe_allow_html=True)
            CorpiScore10Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T3'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T3'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore10Description, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
            CorpiScore10Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T3'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T2'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore10Description, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            CorpiScore12Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T4'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T4'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore12Description, unsafe_allow_html=True)
    elif ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            CorpiScore8Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T6'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T6'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore8Description, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value:
            CorpiScore8Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T6'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T6'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore8Description, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            CorpiScore8Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['T6'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) *int((new_valueb17) * 4)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['T6'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(CorpiScore8Description, unsafe_allow_html=True)
    PitchDescription = f"""
                                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['U2'].value}</td>
                                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['U2'].value:,.0f}</td></tr></table>
                                            """
    st.markdown(PitchDescription, unsafe_allow_html=True)
    if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V2'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V2'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V3'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V3'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V4'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V4'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V5'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V5'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V6'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V6'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V7'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V7'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V8'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V8'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V9'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V9'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V10'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V10'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V11'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V11'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V12'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V12'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V13'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V13'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V14'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V14'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V15'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V15'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V16'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V16'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            TabloFarmanDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['V17'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['V17'].value:,.0f}</td></tr></table>
            """
            st.markdown(TabloFarmanDescription, unsafe_allow_html=True)
    TabloBarghDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['W2'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['W2'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(TabloBarghDescription, unsafe_allow_html=True)
    if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X2'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X2'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X3'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X3'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X4'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X4'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X5'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X5'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X6'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X6'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X7'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X7'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X8'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X8'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X9'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X9'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X10'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X10'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X11'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X11'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X12'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X12'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X13'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X13'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
    elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
        if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X14'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X14'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X15'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X15'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X16'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X16'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
        elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
            UPSDescription = f"""
            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['X17'].value}</td>
            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['X17'].value:,.0f}</td></tr></table>
            """
            st.markdown(UPSDescription, unsafe_allow_html=True)
    # if ws1['B11'].value == ws1['D11'].value:
    #     PhotocellDescription = f"""
    #                     <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;direction:rtl;'>{ws2['Y2'].value}</div>
    #                     """
    #     st.markdown(PhotocellDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value:
        PhotocellDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Y3'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Y3'].value:,.0f}</td></tr></table>
                        """
        st.markdown(PhotocellDescription, unsafe_allow_html=True)
    if ws1['B6'].value == ws1['C6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['D6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z4'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z4'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['E6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z5'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z5'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['F6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z6'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z6'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['G6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z7'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z7'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['H6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z8'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z8'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    elif ws1['B6'].value == ws1['I6'].value:
        CabinDoorShasiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z9'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z9'].value:,.0f}</td></tr></table>
                                """
        st.markdown(CabinDoorShasiDescription, unsafe_allow_html=True)
    LandingShasiDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['Z3'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['Z3'].value:,.0f}</td></tr></table>
                            """
    st.markdown(LandingShasiDescription, unsafe_allow_html=True)
    TravelCableDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AA2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{(int(new_valueb2)) * int(math.ceil(new_valueb8 + 5 + new_valueb9 + 3) + 1)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AA2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(TravelCableDescription, unsafe_allow_html=True)
    ElectrosignalDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AB2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AB2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(ElectrosignalDescription, unsafe_allow_html=True)
    ShalterDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AC2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)*2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AC2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(ShalterDescription, unsafe_allow_html=True)
    OverloadDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AD2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AD2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(OverloadDescription, unsafe_allow_html=True)
    # TunnelLightQuantity = f"""
    #                         <div style='background-color:#f0f0f0;padding:10px;border-radius:5px;display:block;margin-bottom:5px;height:60px;text-align:center;'>{int(new_valueb2 * new_valueb6 * 2)}</div>
    #                         """
    # st.markdown(TunnelLightQuantity, unsafe_allow_html=True)
    TunnelLightDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AE2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2 * new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AE2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(TunnelLightDescription, unsafe_allow_html=True)
    Simafshan6Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(math.ceil(new_valueb8 + new_valueb9 + new_valueb10 + 5))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(Simafshan6Description, unsafe_allow_html=True)
    Simafshan4Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF3'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF3'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(Simafshan4Description, unsafe_allow_html=True)
    if ws1['B6'].value != ws1['C6'].value:
        if new_valueb6 == new_valued6:    
            SimafshanBlueDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF4'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF4'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(SimafshanBlueDescription, unsafe_allow_html=True)
        else:
            SimafshanBlueDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF4'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF4'].value:,.0f}</td></tr></table>
                                    """
            st.markdown(SimafshanBlueDescription, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6:
        SimafshanYellowDescription = f"""
                                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF5'].value}</td>
                                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF5'].value:,.0f}</td></tr></table>
                                                        """
        st.markdown(SimafshanYellowDescription, unsafe_allow_html=True)
    else:
        SimafshanYellowDescription = f"""
                                                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF5'].value}</td>
                                                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF5'].value:,.0f}</td></tr></table>
                                                        """
        st.markdown(SimafshanYellowDescription, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6 or new_valueb6 == new_valuef6:
        SimafshanGreenDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF6'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF6'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanGreenDescription, unsafe_allow_html=True)
    else:
        SimafshanGreenDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF6'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF6'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanGreenDescription, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6 or new_valueb6 == new_valuef6 or new_valueb6 == new_valueg6:
        SimafshanRedDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF7'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF7'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanRedDescription, unsafe_allow_html=True)
    else:
        SimafshanRedDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF7'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF7'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanRedDescription, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6 or new_valueb6 == new_valuef6 or new_valueb6 == new_valueg6 or new_valueb6 == new_valueh6:
        SimafshanBlackDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF8'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF8'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanBlackDescription, unsafe_allow_html=True)
    else:
        SimafshanBlackDescription = f"""
                                                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AF8'].value}</td>
                                                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AF8'].value:,.0f}</td></tr></table>
                                                                """
        st.markdown(SimafshanBlackDescription, unsafe_allow_html=True)
    TunnelLightCableDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AG2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(math.ceil(new_valueb8 + new_valueb9 + new_valueb10 + 5))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AG2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(TunnelLightCableDescription, unsafe_allow_html=True)
    Dockt9Description = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AH2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AH2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(Dockt9Description, unsafe_allow_html=True)
    Dockt3Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AH3'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(math.ceil(((new_valueb8 + new_valueb9 + new_valueb10) / 2) - 1))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AH3'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(Dockt3Description, unsafe_allow_html=True)
    Dockt10Description = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AH4'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AH4'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(Dockt10Description, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6:
        LoleKhortomiFeleziDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AI2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AI2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(LoleKhortomiFeleziDescription, unsafe_allow_html=True)
    else:
        LoleKhortomiFeleziDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AI2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 3}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AI2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(LoleKhortomiFeleziDescription, unsafe_allow_html=True)
    if new_valueb6 == new_valuec6 or new_valueb6 == new_valued6 or new_valueb6 == new_valuee6:
        LoleKhortomiPlasticiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AI3'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AI3'].value:,.0f}</td></tr></table>
                                """
        st.markdown(LoleKhortomiPlasticiDescription, unsafe_allow_html=True)
    else:
        LoleKhortomiPlasticiDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AI3'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 3}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AI3'].value:,.0f}</td></tr></table>
                                """
        st.markdown(LoleKhortomiPlasticiDescription, unsafe_allow_html=True)
    PrizRokarDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AJ2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AJ2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(PrizRokarDescription, unsafe_allow_html=True)
    TabdilKeyDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AK2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AK2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(TabdilKeyDescription, unsafe_allow_html=True)
    GharchiKeyDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AL2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AL2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(GharchiKeyDescription, unsafe_allow_html=True)
    GharchiKeyGhabDescription = f"""
                                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AL3'].value}</td>
                                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AL3'].value:,.0f}</td></tr></table>
                                    """
    st.markdown(GharchiKeyGhabDescription, unsafe_allow_html=True)
    MagnetDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AM2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 5}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AM2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(MagnetDescription, unsafe_allow_html=True)
    NavarChasbDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AN2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 10}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AN2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(NavarChasbDescription, unsafe_allow_html=True)
    BastTravelCableDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AO2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 4}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AO2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(BastTravelCableDescription, unsafe_allow_html=True)
    BastKamarbandiDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AO3'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AO3'].value:,.0f}</td></tr></table>
                            """
    st.markdown(BastKamarbandiDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorFull70LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP2'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP2'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull70LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorFull70RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP3'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull70RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['E13'].value:
        CabinDoorFull70CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP4'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP4'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull70CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorFull80LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP5'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP5'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull80LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorFull80RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP6'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP6'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull80RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['E13'].value:
        CabinDoorFull80CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP7'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP7'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull80CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorFull90LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP8'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP8'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull90LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorFull90RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP9'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP9'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull90RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['E13'].value:
        CabinDoorFull90CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP10'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP10'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull90CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorFull100LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP11'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP11'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull100LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorFull100RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP12'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP12'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull100RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['E13'].value:
        CabinDoorFull100CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP13'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP13'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorFull100CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorNime70LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP14'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP14'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime70LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorNime70RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP15'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP15'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime70RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorNime80LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP16'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP16'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime80LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorNime80RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP17'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP17'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime80RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['C13'].value:
        CabinDoorNime90LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP18'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP18'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime90LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['D13'].value:
        CabinDoorNime90RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AP19'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AP19'].value:,.0f}</td></tr></table>
                """
        st.markdown(CabinDoorNime90RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorFull70LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ2'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ2'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull70LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorFull70RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ3'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ3'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull70RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['E13'].value:
        LandingDoorFull70CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ4'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ4'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull70CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorFull80LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ5'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ5'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull80LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorFull80RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ6'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ6'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull80RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['E13'].value:
        LandingDoorFull80CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ7'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ7'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull80CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorFull90LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ8'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ8'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull90LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorFull90RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ9'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ9'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull90RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['E13'].value:
        LandingDoorFull90CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ10'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ10'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull90CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorFull100LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ11'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ11'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull100LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorFull100RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ12'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ12'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull100RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['C11'].value and ws1['B12'].value == ws1['F12'].value and ws1['B13'].value == ws1['E13'].value:
        LandingDoorFull100CentralDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ13'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ13'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorFull100CentralDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorNime70LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ14'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ14'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime70LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['C12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorNime70RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ15'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ15'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime70RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorNime80LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ16'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ16'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime80LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['D12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorNime80RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ17'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ17'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime80RightDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['C13'].value:
        LandingDoorNime90LeftDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ18'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ18'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime90LeftDescription, unsafe_allow_html=True)
    if ws1['B11'].value == ws1['D11'].value and ws1['B12'].value == ws1['E12'].value and ws1['B13'].value == ws1['D13'].value:
        LandingDoorNime90RightDescription = f"""
                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AQ19'].value}</td>
                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb6)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AQ19'].value:,.0f}</td></tr></table>
                """
        st.markdown(LandingDoorNime90RightDescription, unsafe_allow_html=True)
    # if ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['C19'].value: 
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR2'].value}<br>{ws2['AR19'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR2'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR3'].value}<br>{ws2['AR20'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR3'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR4'].value}<br>{ws2['AR21'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR4'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR5'].value}<br>{ws2['AR22'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR5'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['C19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR6'].value}<br>{ws2['AR23'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR6'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR7'].value}<br>{ws2['AR24'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR7'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR8'].value}<br>{ws2['AR25'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR8'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR9'].value}<br>{ws2['AR26'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR9'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['C5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR10'].value}<br>{ws2['AR27'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR10'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR11'].value}<br>{ws2['AR28'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR11'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR12'].value}<br>{ws2['AR29'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR12'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR13'].value}<br>{ws2['AR30'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR13'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    # elif ws1['B5'].value == ws1['D5'].value and ws1['B19'].value == ws1['D19'].value:
    #     if ws1['B3'].value == ws1['C3'].value or ws1['B3'].value == ws1['D3'].value or ws1['B3'].value == ws1['E3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR14'].value}<br>{ws2['AR31'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR14'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['C4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR15'].value}<br>{ws2['AR32'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR15'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif (ws1['B3'].value == ws1['F3'].value or ws1['B3'].value == ws1['G3'].value) and (ws1['B4'].value == ws1['D4'].value):
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR16'].value}<br>{ws2['AR33'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR16'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    #     elif ws1['B3'].value == ws1['H3'].value or ws1['B3'].value == ws1['I3'].value or ws1['B3'].value == ws1['J3'].value or ws1['B3'].value == ws1['K3'].value or ws1['B3'].value == ws1['L3'].value:
    #         ShasiDescription = f"""
    #         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AR17'].value}<br>{ws2['AR34'].value}</td>
    #         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AR17'].value:,.0f}</td></tr></table>
    #         """
    #         st.markdown(ShasiDescription, unsafe_allow_html=True)
    # if new_valueb14 == new_valuec14 and new_valueb3 == new_valuec3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuec3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valued3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valued3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuee3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuee3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuef3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuef3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueg3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueg3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueh3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueh3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuei3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuei3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuej3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuej3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuek3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuek3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuel3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuel3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuec3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuec3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valued3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valued3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuee3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuee3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuef3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuef3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valueg3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueg3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valueh3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueh3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuei3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuei3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuej3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuej3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuek3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuek3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuel3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuel3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuec3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuec3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valued3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valued3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuee3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuee3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuef3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuef3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueg3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueg3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueh3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valueh3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuei3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuei3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuej3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuej3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuek3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuek3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuel3:
    #     WazneTadolDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int(new_valueb16 + (0.5 * 75 * new_valuel3) - 75)} (کیلوگرم)</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    # if new_valueb14 == new_valuec14 and new_valueb3 == new_valuec3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valued3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuee3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuef3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueg3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueh3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuei3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuej3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuek3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuel3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75)/52)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuec3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valued3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuee3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuef3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valueg3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valueh3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuei3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuej3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuek3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valued14 and new_valueb3 == new_valuel3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75)/58)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuec3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valued3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuee3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuef3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueg3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueh3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuei3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuej3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuek3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    # elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuel3:
    #     WazneDescription = f"""
    #     <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS8'].value}</td>
    #     <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75)/68)}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS8'].value:,.0f}</td></tr></table>
    #     """
    #     st.markdown(WazneDescription, unsafe_allow_html=True)
    if new_valueb14 == new_valuec14 and new_valueb3 == new_valuec3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/52)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valued3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/52)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuee3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/52)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuef3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/52)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueg3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS2'].value}<br>{ws2['AS9'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/52)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS2'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valueh3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuei3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuej3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuek3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuec14 and new_valueb3 == new_valuel3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS3'].value}<br>{ws2['AS10'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS3'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuec3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/58)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valued3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/58)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuee3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/58)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuef3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/58)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valueg3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS4'].value}<br>{ws2['AS11'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/58)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS4'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valueh3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuei3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuej3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuek3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valued14 and new_valueb3 == new_valuel3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS5'].value}<br>{ws2['AS12'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS5'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuec3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuec3) - 75)/68)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valued3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valued3) - 75)/68)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuee3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuee3) - 75)/68)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuef3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuef3) - 75)/68)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueg3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS6'].value}<br>{ws2['AS13'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueg3) - 75)/68)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS6'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valueh3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valueh3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuei3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuei3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuej3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuej3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuek3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuek3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    elif new_valueb14 == new_valuee14 and new_valueb3 == new_valuel3:
        WazneTadolDescription = f"""
        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AS7'].value}<br>{ws2['AS14'].value}</td>
        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * int((new_valueb16 + (0.5 * 75 * new_valuel3) - 75))}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AS7'].value:,.0f}</td></tr></table>
        """
        st.markdown(WazneTadolDescription, unsafe_allow_html=True)
    if new_valueb19 == new_valuec19:    
        HefazFalakeMotorDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AT2'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AT2'].value:,.0f}</td></tr></table>
                        """
        st.markdown(HefazFalakeMotorDescription, unsafe_allow_html=True)
    elif new_valueb19 == new_valued19:    
        HefazFalakeMotorDescription = f"""
                        <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AT5'].value}</td>
                        <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AT5'].value:,.0f}</td></tr></table>
                        """
        st.markdown(HefazFalakeMotorDescription, unsafe_allow_html=True)
    HefazGovernorDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AT3'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AT3'].value:,.0f}</td></tr></table>
                    """
    st.markdown(HefazGovernorDescription, unsafe_allow_html=True)
    HefazFalakeHarzgardDescription = f"""
                    <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AT4'].value}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AT4'].value:,.0f}</td></tr></table>
                    """
    st.markdown(HefazFalakeHarzgardDescription, unsafe_allow_html=True)
    # PhoneDescription = f"""
    #                         <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AU2'].value}</td>
    #                         <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
    #                 <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AU2'].value:,.0f}</td></tr></table>
    #                         """
    # st.markdown(PhoneDescription, unsafe_allow_html=True)
    if new_valueb3 == new_valuec3 or new_valueb3 == new_valued3 or new_valueb3 == new_valuee3 or new_valueb3 == new_valuef3 or new_valueb3 == new_valueg3:
        BatteryDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AV2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 2}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AV2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(BatteryDescription, unsafe_allow_html=True)
    else:
        BatteryDescription = f"""
                                <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AV2'].value}</td>
                                <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2) * 3}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AV2'].value:,.0f}</td></tr></table>
                                """
        st.markdown(BatteryDescription, unsafe_allow_html=True)
    AghlamStandardDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws2['AW2'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;;font-size:1rem;">{int(new_valueb2)}</td>
                    <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{ws3['AW2'].value:,.0f}</td></tr></table>
                            """
    st.markdown(AghlamStandardDescription, unsafe_allow_html=True)

    row_source_names = {
        1: 'MotorDescription',
        2: 'RailT90Description',
        3: 'PoshtBandRailT90Description',
        4: 'RailT70Description',
        5: 'PoshtBandRailT70Description',
        6: 'RailT50Description',
        7: 'PoshtBandRailT50Description',
        8: 'LoghmeRailT90Description',
        9: 'LoghmeRailT70Description',
        10: 'LoghmeRailT50Description',
        11: 'BracketCabinDescription',
        12: 'LeftBracketWazneDescription',
        13: 'RightBracketWazneDescription',
        14: 'Pitch10Description',
        15: 'Pitch12Description',
        16: 'Pitch8Description',
        17: 'LasticDescription',
        18: 'SimBoxelScore10Description',
        19: 'SimBoxelScore8Description',
        20: 'SimBoxelScore12Description',
        21: 'SimBoxelScore6Description',
        22: 'GholabBoxelScore10Description',
        23: 'GholabBoxelScore8Description',
        24: 'GholabBoxelScore13Description',
        25: 'FalakeHarzgardDescription',
        26: 'GovernorUpDescription',
        27: 'LeftDoorGhoflDescription',
        28: 'RightDoorGhoflDescription',
        29: 'DictatorDescription',
        30: 'DoorKeyDescription',
        31: 'PolyortanDescription',
        32: 'BufferDescription',
        33: 'RoghandanDescription',
        34: 'AshkiDescription',
        35: 'CorpiScore6Description',
        36: 'CorpiScore10Description',
        37: 'CorpiScore12Description',
        38: 'CorpiScore8Description',
        39: 'PitchDescription',
        40: 'TabloFarmanDescription',
        41: 'TabloBarghDescription',
        42: 'UPSDescription',
        43: 'PhotocellDescription',
        44: 'CabinDoorShasiDescription',
        45: 'LandingShasiDescription',
        46: 'TravelCableDescription',
        47: 'ElectrosignalDescription',
        48: 'ShalterDescription',
        49: 'OverloadDescription',
        50: 'TunnelLightDescription',
        51: 'Simafshan6Description',
        52: 'Simafshan4Description',
        53: 'SimafshanBlueDescription',
        54: 'SimafshanYellowDescription',
        55: 'SimafshanGreenDescription',
        56: 'SimafshanRedDescription',
        57: 'SimafshanBlackDescription',
        58: 'TunnelLightCableDescription',
        59: 'Dockt9Description',
        60: 'Dockt3Description',
        61: 'Dockt10Description',
        62: 'LoleKhortomiFeleziDescription',
        63: 'LoleKhortomiPlasticiDescription',
        64: 'PrizRokarDescription',
        65: 'TabdilKeyDescription',
        66: 'GharchiKeyDescription',
        67: 'GharchiKeyGhabDescription',
        68: 'MagnetDescription',
        69: 'NavarChasbDescription',
        70: 'BastTravelCableDescription',
        71: 'BastKamarbandiDescription',
        72: 'CabinDoorFull70LeftDescription',
        73: 'CabinDoorFull70RightDescription',
        74: 'CabinDoorFull70CentralDescription',
        75: 'CabinDoorFull80LeftDescription',
        76: 'CabinDoorFull80RightDescription',
        77: 'CabinDoorFull80CentralDescription',
        78: 'CabinDoorFull90LeftDescription',
        79: 'CabinDoorFull90RightDescription',
        80: 'CabinDoorFull90CentralDescription',
        81: 'CabinDoorFull100LeftDescription',
        82: 'CabinDoorFull100RightDescription',
        83: 'CabinDoorFull100CentralDescription',
        84: 'CabinDoorNime70LeftDescription',
        85: 'CabinDoorNime70RightDescription',
        86: 'CabinDoorNime80LeftDescription',
        87: 'CabinDoorNime80RightDescription',
        88: 'CabinDoorNime90LeftDescription',
        89: 'CabinDoorNime90RightDescription',
        90: 'LandingDoorFull70LeftDescription',
        91: 'LandingDoorFull70RightDescription',
        92: 'LandingDoorFull70CentralDescription',
        93: 'LandingDoorFull80LeftDescription',
        94: 'LandingDoorFull80RightDescription',
        95: 'LandingDoorFull80CentralDescription',
        96: 'LandingDoorFull90LeftDescription',
        97: 'LandingDoorFull90RightDescription',
        98: 'LandingDoorFull90CentralDescription',
        99: 'LandingDoorFull100LeftDescription',
        100: 'LandingDoorFull100RightDescription',
        101: 'LandingDoorFull100CentralDescription',
        102: 'LandingDoorNime70LeftDescription',
        103: 'LandingDoorNime70RightDescription',
        104: 'LandingDoorNime80LeftDescription',
        105: 'LandingDoorNime80RightDescription',
        106: 'LandingDoorNime90LeftDescription',
        107: 'LandingDoorNime90RightDescription',
        108: 'ShasiDescription',
        109: 'WazneTadolDescription',
        110: 'HefazFalakeMotorDescription',
        111: 'HefazGovernorDescription',
        112: 'HefazFalakeHarzgardDescription',
        113: 'BatteryDescription',
        114: 'AghlamStandardDescription'
    }    
    
    Products = {}
    Quantities = {}
    Unit_Prices = {}
    for i in range(1, 115):
        try:
            source_name = row_source_names.get(i)
            if not source_name:
                continue
            row_html = globals().get(source_name)
            if not row_html or not isinstance(row_html, str):
                continue
            soup = BeautifulSoup(row_html, "html.parser")
            tds = soup.find_all("td")
            if len(tds) < 3:
                continue
            Quantity = int(tds[1].text.strip().replace(",", ""))
            Unit_Price = int(tds[2].text.strip().replace(",", ""))
            Quantities[i] = Quantity
            Unit_Prices[i] = Unit_Price
            Products[i] = Quantity * Unit_Price
            print(f"Row {i} => {Quantity} * {Unit_Price} = {Products[i]}")   
        except Exception as e:
            continue
    Total = sum(Products.values())
    print("Total:", Total)
    TotalPriceDescription = f"""
                            <table style="width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;height:60px;direction:rtl;background-color:#f0f0f0;border-radius:5px;"><tr><td style="width:80%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-right-radius:5px;border-bottom-right-radius:5px;">{ws3['BA1'].value}</td>
                            <td style="width:10%;border:1px solid #ddd;padding:10px;white-space:nowrap;text-align:center;vertical-align:middle;border-top-left-radius:5px;border-bottom-left-radius:5px;;font-size:1rem;">{Total:,.0f}</td></tr></table>
                            """
    st.markdown(TotalPriceDescription, unsafe_allow_html=True)








